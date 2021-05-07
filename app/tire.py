from tkinter.constants import RADIOBUTTON, TRUE
from typing import Text
from helium import*
from bs4 import BeautifulSoup
import pandas as pd
import os
import time
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
import datetime
from selenium.webdriver.common.by import By
from openpyxl import Workbook
from openpyxl import load_workbook
import json
from pathlib import Path
import http.client

class GeneticRisk():

    def __init__(self):
        pass
    def execute(self,test_sheet,case_num):
        self.test_sheet = test_sheet
        user_dict = self.read_data()
        cases = user_dict.keys()
        for self.case_iter,self.case_num in enumerate(cases):
            #print(user_dict.keys())
            if self.case_num == case_num:
                print(user_dict[self.case_num])
                return self.calculate(user_dict[self.case_num])
    




    def read_data(self):
        data_f = pd.read_excel(self.test_sheet)

        chl = 0
        for inx,rows in data_f.iterrows():
            if rows['Sub-Feature'] == "Relative-relation":
                chl = inx
        param_set_1 = data_f.iloc[1:chl,3:]

        param_set_2 = data_f.iloc[chl:,3:]
        cols = param_set_2.iloc[:,1:].columns
        relative_dict = {}
        user_dict = {}
        for col in cols:
            relative_dict[col] = {}
            user_dict[col] = {}




        num_rel = 0
        for inx,row in param_set_2.iterrows():

            if row['Sub-Feature'] == "external factor":
                break

            if row['Sub-Feature'] == "Relative-relation":
                num_rel = num_rel + 1
                
                for idx,col in enumerate(cols):

                    relative_dict[col][num_rel]= {}
                    relative_dict[col][num_rel]['name'] = row[col]
            elif row['Sub-Feature'] == "Relative-current age/age of death":
                for idx,col in enumerate(cols):
                    relative_dict[col][num_rel]['cur_age'] = row[col]
            elif row['Sub-Feature'] == "Relative-bc status":

                for idx,col in enumerate(cols):
                    relative_dict[col][num_rel]['bc_status'] = row[col]
            elif row['Sub-Feature'] == "Relative-bilateral":

                for idx,col in enumerate(cols):
                    relative_dict[col][num_rel]['bilateral_status'] = row[col]
            elif row['Sub-Feature'] == "Relative-bc status":

                for idx,col in enumerate(cols):
                    relative_dict[col][num_rel]['bc_status'] = row[col]
            elif row['Sub-Feature'] == "Relative-bc_onset":

                for idx,col in enumerate(cols):
                    relative_dict[col][num_rel]['bc_onset'] = row[col]
            elif row['Sub-Feature'] == "Relative-oc_status":

                for idx,col in enumerate(cols):
                    relative_dict[col][num_rel]['oc_status'] = row[col]
            elif row['Sub-Feature'] == "Relative-oc_onset":

                for idx,col in enumerate(cols):
                    relative_dict[col][num_rel]['oc_onset'] = row[col]

            elif row['Sub-Feature'] == "Relative-brca_gene":

                for idx,col in enumerate(cols):
                    relative_dict[col][num_rel]['brca_gene'] = row[col]

            else:
                pass

        param_set_1 = param_set_1.fillna(0)
        for inx,row in param_set_1.iterrows():
            if row['Sub-Feature'] == "Weight":
                for col in cols:
                    user_dict[col]['weight'] = row[col]
            elif row['Sub-Feature'] == "Name":
                for col in cols:
                    user_dict[col]['name'] = row[col]
            elif row['Sub-Feature'] == "Accompanied by":
                for col in cols:
                    user_dict[col]['accompanied_by'] = row[col]
            elif row['Sub-Feature'] == "Address":
                for col in cols:
                    user_dict[col]['address'] = row[col]
            elif row['Sub-Feature'] == "Date":
                for col in cols:
                    user_dict[col]['date'] = row[col]
            elif row['Sub-Feature'] == "Age":
                for col in cols:
                    user_dict[col]['age'] = row[col]
            elif row['Sub-Feature'] == "Issues":
                for col in cols:
                    user_dict[col]['issues'] = row[col]
            elif row['Sub-Feature'] == "Ethinicity":
                for col in cols:
                    user_dict[col]['ethinicity'] = row[col]
            elif row['Sub-Feature'] == "Height":
                for col in cols:
                    user_dict[col]['height'] = row[col]
            elif row['Sub-Feature'] == "Breast Biopsy":
                for col in cols:
                    user_dict[col]['biopsy'] = row[col]
            elif row['Sub-Feature'] == "Parity":
                for col in cols:
                    user_dict[col]['parity'] = row[col]
            elif row['Sub-Feature'] == "Breast/Mamographic Density":
                for col in cols:
                    user_dict[col]['md'] = row[col]
            elif row['Sub-Feature'] == "Age @ Menarche":
                for col in cols:
                    user_dict[col]['menarche'] = row[col]
            elif row['Sub-Feature'] == "Age at first live birth of a child":
                for col in cols:
                    user_dict[col]['flb'] = row[col]
            elif row['Sub-Feature'] == "Age @ Menopause":
                for col in cols:
                    user_dict[col]['menopause'] = row[col]

            elif row['Sub-Feature'] == "Menopause Status":
                for col in cols:
                    user_dict[col]['menopause_status'] = row[col]

            elif row['Sub-Feature'] == "Alcohol Intake":
                for col in cols:
                    user_dict[col]['alcohol'] = row[col]
            elif row['Sub-Feature'] == "OC-Pills":
                for col in cols:
                    user_dict[col]['oc_pills'] = row[col]
            elif row['Sub-Feature'] == "BRCA Status":
                for col in cols:
                    user_dict[col]['brca'] = row[col]
            elif row['Sub-Feature'] == "Ovarian Cancer":
                for col in cols:
                    user_dict[col]['oc'] = row[col]
            elif row['Sub-Feature'] == "Ovarian Cancer@Age":
                for col in cols:
                    user_dict[col]['oc_age'] = row[col]
            elif row['Sub-Feature'] == "Jewish":
                for col in cols:
                    user_dict[col]['jewish'] = row[col]

            elif row['Sub-Feature'] == "HRT Status":
                for col in cols:
                    user_dict[col]['hrt_status'] = row[col]
            elif row['Sub-Feature'] == "HRT(if Current) Length of use":
                for col in cols:
                    user_dict[col]['hrt_cur_len'] = row[col]
            elif row['Sub-Feature'] == "HRT(if Current) Intended Length of use":
                for col in cols:
                    user_dict[col]['hrt_int_len'] = row[col]
            elif row['Sub-Feature'] == "HRT(if Current) Type":
                for col in cols:
                    user_dict[col]['hrt_cur_type'] = row[col]
            elif row['Sub-Feature'] == "HRT(if Past) Last used year back":
                for col in cols:
                    user_dict[col]['hrt_past_use'] = row[col]
            elif row['Sub-Feature'] == "HRT(if Past) Past Length of use":
                for col in cols:
                    user_dict[col]['hrt_past_len'] = row[col]
            elif row['Sub-Feature'] == "HRT(if Past) Type":
                for col in cols:
                    user_dict[col]['hrt_past_type'] = row[col]
            else:
                pass
        for col in cols:
            user_dict[col]['FH'] = relative_dict[col]
        return user_dict


    def calculate(self,user_dict):


        options = Options()
        # options.binary_location = r"C:\Users\amisr\Downloads\03-01-2021\Genetic_automation\chromium-87-0-4246\chrome-win\chrome.exe"
        options.binary_location = r"D:\Win_x64_843846_chrome-win\chrome-win\chrome.exe"
        options.add_argument("--disable-infobars")
        options.add_argument("start-maximized")
        options.add_argument("--disable-extensions")

        # Pass the argument 1 to allow and 2 to block
        options.add_experimental_option("prefs", { 
            "profile.default_content_setting_values.notifications": 1 
        })

        #driver = webdriver.Chrome(executable_path="chromedriver.exe", chrome_options=options)
        #driverh =  set_driver(driver)
        #driver.get('https://www.facebook.com/YoVille')
        f = open('Value-json/logic_activation.json') 
        activation = json.load(f)
        if activation['headless'] == "active":

            head_param = True
        else:
            head_param = False
        self.driver= start_chrome('https://ibis-risk-calculator.magview.com/',options=options,headless=head_param)
     

        
        write(user_dict['age'],into = 'YOUR AGE')
        press(ENTER)
        press(TAB)
        press(TAB)
        

        press(TAB)
        press(ARROW_LEFT)
        press(TAB)
        write(str(user_dict['height']))
        press(ENTER)
       
        press(TAB)
        write(str(user_dict['weight']))
        press(ENTER)                                                                                     
       
        press(TAB)
        write(str(user_dict['menarche']))
        press(ENTER)
        
        click('NEXT')
        press(TAB)
        press(ARROW_LEFT)
        if user_dict.get('parity') is not None:
            if user_dict['parity'] == 0:
                press(ARROW_LEFT)
                press(TAB)
                press(ARROW_LEFT)
                
            else:
                press(ARROW_LEFT)
                press(ARROW_LEFT)
                press(TAB)
                write(user_dict['flb'])
                press(ENTER)
                press(TAB)
                press(TAB)
                
                press(TAB)
                press(ARROW_LEFT)


        else:
            press(TAB)
            press(ARROW_LEFT)

        if user_dict.get('menopause_status') is not None:
            if user_dict['menopause_status'] == 'Postmenopausal':
                press(ARROW_LEFT)
                press(TAB)
                write(user_dict['menopause'])
                press(ENTER)
                press(TAB)
                press(TAB)
                

                press(TAB)
                press(ARROW_LEFT)
            # elif user_dict['menopause_status'] == 'Perimenopausal':
            #     press(ARROW_LEFT)
            #     press(ARROW_LEFT)
            #     press(TAB)
            #     press(ARROW_LEFT)

            elif user_dict['menopause_status'] == 'Premenopausal' or user_dict['menopause_status'] == 'Perimenopausal':
                press(ARROW_LEFT)
                press(ARROW_LEFT)
                press(ARROW_LEFT)
                press(TAB)
                press(ARROW_LEFT)
            else:
                #print("No menopause status")
                press(TAB)
                press(ARROW_LEFT)
        
            




        else:
            press(TAB)
            press(ARROW_LEFT)


        if user_dict.get('hrt_status') is not None:
            if user_dict['hrt_status'] == 'Never':
                press(ARROW_LEFT)
                press(TAB)
                press(ARROW_LEFT)
            elif user_dict['hrt_status'] == 'Current User':
                #print("in the current")
                press(TAB)
                write(user_dict['hrt_cur_len'])
                press(ENTER)
                press(TAB)
  
                
   
                write(user_dict['hrt_int_len'])
                press(ENTER)
                press(TAB)
 
                if user_dict['hrt_cur_type'] == 'Combined':
                    press(TAB)
                    press(ARROW_LEFT)
                else:
                    press(ARROW_LEFT)
                    press(TAB)
                    press(ARROW_LEFT)






            elif user_dict['hrt_status'] == 'Less than 5 years ago':
                press(ARROW_LEFT)
                press(ARROW_LEFT)
                press(TAB)
                write(user_dict['hrt_past_use'])
                press(ENTER)
                press(TAB)
                write(user_dict['hrt_past_len'])
                press(ENTER)
                press(TAB)
                if user_dict['hrt_past_type'] == 'Combined':
                    press(TAB)
                    press(ARROW_LEFT)
                else:
                    press(ARROW_UP)
                    press(TAB)
                    press(ARROW_LEFT)
            elif user_dict['hrt_status'] == '5+ years ago':
                press(ARROW_LEFT)
                press(ARROW_LEFT)
                press(ARROW_LEFT)
                press(TAB)
                press(ARROW_LEFT)

                
                

            else:
                #print("No hrt status")
                press(ARROW_LEFT)
                press(TAB)
                press(ARROW_LEFT)
        
            




        else:
            press(ARROW_LEFT)
            press(TAB)
            press(ARROW_LEFT)

        if user_dict.get('brca') is not None:
            if user_dict['brca'] == 'brca2':
                press(TAB)
                
            elif user_dict['brca'] == 'brca1':
                press(ARROW_LEFT)
                press(TAB)
                
            elif user_dict['brca'] == 'Negative':
                press(ARROW_LEFT)
                press(ARROW_LEFT)
                press(TAB)
                
            else:
                press(ARROW_LEFT)
                press(ARROW_LEFT)
                press(ARROW_LEFT)
                press(TAB)
                
                
        



        else:
            press(ARROW_LEFT)
            press(ARROW_LEFT)
            press(ARROW_LEFT)
            press(TAB)
            

        if user_dict.get('biopsy') is not None:
            if user_dict['biopsy']=='No prior biopsy / No proliferative disease':
                press(TAB)
                press(TAB)
                press(TAB)
                press(ARROW_LEFT)
            elif user_dict['biopsy']=='Prior biopsy, result unknown':
                press(ARROW_RIGHT)
                press(TAB)
                press(TAB)
                press(TAB)
                press(ARROW_LEFT)
            elif user_dict['biopsy']=='Hyperplasia (not atypia)':
                press(ARROW_RIGHT)
                press(ARROW_RIGHT)
                press(TAB)
                press(TAB)
                press(TAB)
                press(ARROW_LEFT)
            elif user_dict['biopsy']=='Atypical Hyperplasia':
                click('Atypical Hyperplasia')   
                press(TAB)
                press(TAB)
                press(ARROW_LEFT)
            elif user_dict['biopsy']=='Lobular Carcinoma in Situ (LCIS)':
                click('Lobular Carcinoma in Situ (LCIS)') 
                press(TAB)
                press(ARROW_LEFT)
                
            else:
                
                press(TAB)
                press(TAB)
                press(TAB)
                press(ARROW_LEFT)

        else:

            press(TAB)
            press(TAB)
            press(TAB)
            press(ARROW_LEFT)



        if user_dict.get('oc') is not None:
            if user_dict['oc'] == "I don't know":
                press(TAB)
                
            elif user_dict['oc'] == 'No':
                press(ARROW_LEFT)
                press(TAB)
     
            elif user_dict['oc'] == 'Yes':
                press(ARROW_LEFT)
                press(ARROW_LEFT)
                press(TAB)
                write(user_dict['oc_age'])
                press(ENTER)
                press(TAB)
                press(TAB)
                press(TAB)

            else:
                press(TAB)

                
        



        else:
 
            press(TAB)
        

        if user_dict.get('md') is not None:
            if user_dict['md'] == 'Almost entirely fatty':
                                
                press(TAB)
                press(ENTER)
                press(ARROW_UP)
                press(ARROW_UP)
                press(ARROW_UP)
                press(ARROW_UP)
                press(TAB)
                press(ARROW_LEFT)
            elif user_dict['md'] == 'Scattered fibroglandular density':
                press(TAB)
                press(ENTER)
                press(ARROW_UP)
                press(ARROW_UP)
                press(ARROW_UP)
                press(TAB)
                press(ARROW_LEFT)
            elif user_dict['md'] == 'Heterogenously dense':
                press(TAB)
                press(ENTER)
                press(ARROW_UP)
                press(ARROW_UP)

                press(TAB)
                press(ARROW_LEFT)
            elif user_dict['md'] == 'Extremely dense':
                press(TAB)
                press(ENTER)
                press(ARROW_UP)

                press(TAB)
                press(ARROW_LEFT)
            elif user_dict['md'] == "I don't know":
                press(TAB)
                press(ENTER)

                press(TAB)
                press(ARROW_LEFT)
            else:
                press(TAB)
                press(ENTER)

                press(TAB)
                press(ARROW_LEFT)
            
        




        else:
            press(ARROW_LEFT)
            press(TAB)
            press(ARROW_LEFT)


        if user_dict.get('jewish') is not None:
            if user_dict['jewish'] == 'No':
                press(ARROW_LEFT)
                click('NEXT')
                
            else:
                press(ARROW_LEFT)
                press(ARROW_LEFT)
                click('NEXT')
        else:
            click('NEXT')

        relative_dict = user_dict['FH']


      

        for inx,relative in relative_dict.items():
            num_tab_max = 13
            seq_dict = {'Mother' : 0,
                        'Father': 1,
                        'Sister' : 2,
                        'Daughter' : 3,
                        'Brother' : 4,
                        'Paternal Grandmother':5,
                        'Paternal Aunt':6,
                        'Paternal Half Sister':7,
                        "Paternal Uncle's Daughter":8,
                        'Maternal Grandmother':9,
                        'Maternal Aunt':10,
                        'Maternal Half Sister':11,
                        "Maternal Uncle's Daughter":12}
            if relative['name']!=relative['name']:
                pass
            else:
                to_click = str(relative['name']).upper()
                
                click(Link(to_click))
                
                
                
                for i in range(13-int(seq_dict[relative['name']])):
                    
                    press(TAB)



                


                if relative['name'] == 'Mother':
                    if relative['bc_status'] == 'Yes':
                        press(ARROW_LEFT)
                        press(ARROW_LEFT)
                        press(TAB)
                        write(relative['cur_age'])
                        press(ENTER)
                        press(TAB)
                        press(TAB) 
                        press(TAB)
                        press(ARROW_LEFT)
                        
                        if relative['bilateral_status'] == 'Yes':
                            press(ARROW_LEFT)
                            press(ARROW_LEFT)
                            press(TAB)
                            write(relative['bc_onset'])
                            press(ENTER)
                            press(TAB)
                            press(TAB)
                            press(TAB)

                            
                        else:
                            press(ARROW_LEFT)
                            press(TAB)





                    else:
                        press(ARROW_LEFT)
                        press(TAB)
                        write(relative['cur_age'])
                        press(ENTER)
                        press(TAB)
                        press(TAB)
                        press(TAB)
                    

                    if relative['oc_status'] == 'Yes':
                        press(ARROW_LEFT)
                        press(ARROW_LEFT)

                        press(TAB)
                        write(relative['oc_onset'])
                        press(ENTER)
                        press(TAB)
                        press(TAB)
                        press(TAB)
                    else:
                        press(ARROW_LEFT)
                        press(TAB)

                    if relative.get('brca_gene') is not None:
                        if relative['brca_gene'] == 'Negative':
                            press(ARROW_RIGHT)
                        elif relative['brca_gene'] == 'BRCA1':
                            press(ARROW_RIGHT)
                            press(ARROW_RIGHT)
                        elif relative['brca_gene'] == 'BRCA2':
                            press(ARROW_RIGHT)
                            press(ARROW_RIGHT)
                            press(ARROW_RIGHT)
                        else:
                            pass
                        
                    else:
                        pass

                
                elif relative['name'] == 'Father':
                    if relative['bc_status'] == 'Yes':
                        press(ARROW_LEFT)
                        press(ARROW_LEFT)
                        press(TAB)
                        write(relative['cur_age'])
                        press(ENTER)
                        press(TAB)
                        press(TAB)
                        press(TAB)
                        





                    else:
                        press(ARROW_LEFT)
                        press(TAB)
                        write(relative['cur_age'])
                        press(ENTER)
                        press(TAB)
                        press(TAB)
                        press(TAB)
                    


                    if relative.get('brca_gene') is not None:
                        if relative['brca_gene'] == 'Negative':
                            press(ARROW_RIGHT)
                        elif relative['brca_gene'] == 'BRCA1':
                            press(ARROW_RIGHT)
                            press(ARROW_RIGHT)
                        elif relative['brca_gene'] == 'BRCA2':
                            press(ARROW_RIGHT)
                            press(ARROW_RIGHT)
                            press(ARROW_RIGHT)
                        else:
                            pass
                        
                    else:
                        pass

    



                
                elif relative['name'] == 'Sister':
                    write('1')
                    press(ENTER)
                    press(TAB)
                    press(TAB)
                    click('GO TO DETAILS')

                    

                    to_click = str(relative['name']).upper()
                    
                    click(Link(to_click))

                    for i in range(13-int(seq_dict[relative['name']])):
                    
                        press(TAB)
                 
                    if relative['bc_status'] == 'Yes':
                        press(ARROW_LEFT)
                        press(ARROW_LEFT)
                        press(ARROW_LEFT)
                        press(TAB)
                        write(relative['cur_age'])
                        press(ENTER)
                        press(TAB)
                        press(TAB) 
                        press(TAB)
                        press(ARROW_LEFT)
                        
                        if relative['bilateral_status'] == 'Yes':
                            press(ARROW_LEFT)
                            press(ARROW_LEFT)
                            press(TAB)
                             
                            write(relative['bc_onset'])
                            press(ENTER)
                            press(TAB)
                            press(TAB)
                            press(TAB)

                            
                        else:
                            press(ARROW_LEFT)
                            press(TAB)





                    else:
                        press(ARROW_LEFT)
                        press(ARROW_LEFT)
                        press(TAB)
                        write(relative['cur_age'])
                        press(ENTER)
                        press(TAB)
                        press(TAB)
                        press(TAB)
                    

                    if relative['oc_status'] == 'Yes':
                        press(ARROW_LEFT)
                        press(ARROW_LEFT)

                        press(TAB)
                        write(relative['oc_onset'])
                        press(ENTER)
                        press(TAB)
                        press(TAB)
                        press(TAB)
                    else:
                        press(ARROW_LEFT)
                        press(TAB)

                    if relative.get('brca_gene') is not None:
                        if relative['brca_gene'] == 'Negative':
                            press(ARROW_RIGHT)
                        elif relative['brca_gene'] == 'BRCA1':
                            press(ARROW_RIGHT)
                            press(ARROW_RIGHT)
                        elif relative['brca_gene'] == 'BRCA2':
                            press(ARROW_RIGHT)
                            press(ARROW_RIGHT)
                            press(ARROW_RIGHT)
                        else:
                            pass
                        
                    else:
                        pass
                elif relative['name'] == 'Daughter':
                    write('1')
                    press(ENTER)
                    press(TAB)
                    press(TAB)
                    click('GO TO DETAILS')

                    

                    to_click = str(relative['name']).upper()
                    
                    click(Link(to_click))

                    for i in range(13-int(seq_dict[relative['name']])):
                    
                        press(TAB)
                 
                    if relative['bc_status'] == 'Yes':
                        press(ARROW_LEFT)
                        press(ARROW_LEFT)
                        press(ARROW_LEFT)
                        press(TAB)
                        write(relative['cur_age'])
                        press(ENTER)
                        press(TAB)
                        press(TAB) 
                        press(TAB)
                        press(ARROW_LEFT)
                        
                        if relative['bilateral_status'] == 'Yes':
                            press(ARROW_LEFT)
                            press(ARROW_LEFT)
                            press(TAB)
                             
                            write(relative['bc_onset'])
                            press(ENTER)
                            press(TAB)
                            press(TAB)
                            press(TAB)

                            
                        else:
                            press(ARROW_LEFT)
                            press(TAB)





                    else:
                        press(ARROW_LEFT)
                        press(ARROW_LEFT)
                        press(TAB)
                        write(relative['cur_age'])
                        press(ENTER)
                        press(TAB)
                        press(TAB)
                        press(TAB)
                    

                    if relative['oc_status'] == 'Yes':
                        press(ARROW_LEFT)
                        press(ARROW_LEFT)

                        press(TAB)
                        write(relative['oc_onset'])
                        press(ENTER)
                        press(TAB)
                        press(TAB)
                        press(TAB)
                    else:
                        press(ARROW_LEFT)
                        press(TAB)

                    if relative.get('brca_gene') is not None:
                        if relative['brca_gene'] == 'Negative':
                            press(ARROW_RIGHT)
                        elif relative['brca_gene'] == 'BRCA1':
                            press(ARROW_RIGHT)
                            press(ARROW_RIGHT)
                        elif relative['brca_gene'] == 'BRCA2':
                            press(ARROW_RIGHT)
                            press(ARROW_RIGHT)
                            press(ARROW_RIGHT)
                        else:
                            pass
                        
                    else:
                        pass

                elif relative['name'] == "Paternal Uncle's Daughter":
                    write('1')
                    press(ENTER)
                    press(TAB)
                    press(TAB)
                    click('GO TO DETAILS')

                    

                    to_click = str(relative['name']).upper()
                    
                    click(Link(to_click))

                    for i in range(13-int(seq_dict[relative['name']])):
                    
                        press(TAB)
                 
                    if relative['bc_status'] == 'Yes':
                        press(ARROW_LEFT)
                        press(ARROW_LEFT)
                        press(ARROW_LEFT)
                        press(TAB)
                        write(relative['cur_age'])
                        press(ENTER)
                        press(TAB)
                        press(TAB) 
                        press(TAB)
                        press(ARROW_LEFT)
                        
                        if relative['bilateral_status'] == 'Yes':
                            press(ARROW_LEFT)
                            press(ARROW_LEFT)
                            press(TAB)
                             
                            write(relative['bc_onset'])
                            press(ENTER)
                            press(TAB)
                            press(TAB)
                            press(TAB)

                            
                        else:
                            press(ARROW_LEFT)
                            press(TAB)





                    else:
                        press(ARROW_LEFT)
                        press(ARROW_LEFT)
                        press(TAB)
                        write(relative['cur_age'])
                        press(ENTER)
                        press(TAB)
                        press(TAB)
                        press(TAB)
                    

                    if relative['oc_status'] == 'Yes':
                        press(ARROW_LEFT)
                        press(ARROW_LEFT)

                        press(TAB)
                        write(relative['oc_onset'])
                        press(ENTER)
                        press(TAB)
                        press(TAB)
                        press(TAB)
                    else:
                        press(ARROW_LEFT)
                        press(TAB)

                    if relative.get('brca_gene') is not None:
                        if relative['brca_gene'] == 'Negative':
                            press(ARROW_RIGHT)
                        elif relative['brca_gene'] == 'BRCA1':
                            press(ARROW_RIGHT)
                            press(ARROW_RIGHT)
                        elif relative['brca_gene'] == 'BRCA2':
                            press(ARROW_RIGHT)
                            press(ARROW_RIGHT)
                            press(ARROW_RIGHT)
                        else:
                            pass
                        
                    else:
                        pass

                elif relative['name'] == "Maternal Uncle's Daughter":
                    write('1')
                    press(ENTER)
                    press(TAB)
                    press(TAB)
                    click('GO TO DETAILS')

                    

                    to_click = str(relative['name']).upper()
                    
                    click(Link(to_click))

                    for i in range(13-int(seq_dict[relative['name']])):
                    
                        press(TAB)
                 
                    if relative['bc_status'] == 'Yes':
                        press(ARROW_LEFT)
                        press(ARROW_LEFT)
                        press(ARROW_LEFT)
                        press(TAB)
                        write(relative['cur_age'])
                        press(ENTER)
                        press(TAB)
                        press(TAB) 
                        press(TAB)
                        press(ARROW_LEFT)
                        
                        if relative['bilateral_status'] == 'Yes':
                            press(ARROW_LEFT)
                            press(ARROW_LEFT)
                            press(TAB)
                             
                            write(relative['bc_onset'])
                            press(ENTER)
                            press(TAB)
                            press(TAB)
                            press(TAB)

                            
                        else:
                            press(ARROW_LEFT)
                            press(TAB)





                    else:
                        press(ARROW_LEFT)
                        press(ARROW_LEFT)
                        press(TAB)
                        write(relative['cur_age'])
                        press(ENTER)
                        press(TAB)
                        press(TAB)
                        press(TAB)
                    

                    if relative['oc_status'] == 'Yes':
                        press(ARROW_LEFT)
                        press(ARROW_LEFT)

                        press(TAB)
                        write(relative['oc_onset'])
                        press(ENTER)
                        press(TAB)
                        press(TAB)
                        press(TAB)
                    else:
                        press(ARROW_LEFT)
                        press(TAB)

                    if relative.get('brca_gene') is not None:
                        if relative['brca_gene'] == 'Negative':
                            press(ARROW_RIGHT)
                        elif relative['brca_gene'] == 'BRCA1':
                            press(ARROW_RIGHT)
                            press(ARROW_RIGHT)
                        elif relative['brca_gene'] == 'BRCA2':
                            press(ARROW_RIGHT)
                            press(ARROW_RIGHT)
                            press(ARROW_RIGHT)
                        else:
                            pass
                        
                    else:
                        pass



                elif relative['name'] == 'Brother':
                    if relative['bc_status'] == 'Yes':
                        press(ARROW_LEFT)
                        press(ARROW_LEFT)
                        press(TAB)
                        write(relative['cur_age'])
                        press(ENTER)
                        press(TAB)
                        press(TAB)
                      
                        





                    else:
                        press(ARROW_LEFT)
                        press(TAB)
                        write(relative['cur_age'])
                        press(ENTER)
                        press(TAB)
                        press(TAB)


                elif relative['name'] == 'Paternal Grandmother':
                    if relative['bc_status'] == 'Yes':
                        press(ARROW_LEFT)
                        press(ARROW_LEFT)
                        press(TAB)
                        write(relative['cur_age'])
                        press(ENTER)
                        press(TAB)
                        press(TAB) 
                        press(TAB)
                        press(ARROW_LEFT)
                        
                        if relative['bilateral_status'] == 'Yes':
                            press(ARROW_LEFT)
                            press(ARROW_LEFT)
                            press(TAB)
                            write(relative['bc_onset'])
                            press(ENTER)
                            press(TAB)
                            press(TAB)
                            press(TAB)

                            
                        else:
                            press(ARROW_LEFT)
                            press(TAB)





                    else:
                        press(ARROW_LEFT)
                        press(TAB)
                        write(relative['cur_age'])
                        press(ENTER)
                        press(TAB)
                        press(TAB)
                        press(TAB)
                    

                    if relative['oc_status'] == 'Yes':
                        press(ARROW_LEFT)
                        press(ARROW_LEFT)

                        press(TAB)
                        write(relative['oc_onset'])
                        press(ENTER)
                        press(TAB)
                        press(TAB)
                        press(TAB)
                    else:
                        press(ARROW_LEFT)
                        press(TAB)

                    if relative.get('brca_gene') is not None:
                        if relative['brca_gene'] == 'Negative':
                            press(ARROW_RIGHT)
                        elif relative['brca_gene'] == 'BRCA1':
                            press(ARROW_RIGHT)
                            press(ARROW_RIGHT)
                        elif relative['brca_gene'] == 'BRCA2':
                            press(ARROW_RIGHT)
                            press(ARROW_RIGHT)
                            press(ARROW_RIGHT)
                        else:
                            pass
                        
                    else:
                        pass

                elif relative['name'] == 'Maternal Grandmother':
                    if relative['bc_status'] == 'Yes':
                        press(ARROW_LEFT)
                        press(ARROW_LEFT)
                        press(TAB)
                        write(relative['cur_age'])
                        press(ENTER)
                        press(TAB)
                        press(TAB) 
                        press(TAB)
                        press(ARROW_LEFT)
                        
                        if relative['bilateral_status'] == 'Yes':
                            press(ARROW_LEFT)
                            press(ARROW_LEFT)
                            press(TAB)
                            write(relative['bc_onset'])
                            press(ENTER)
                            press(TAB)
                            press(TAB)
                            press(TAB)

                            
                        else:
                            press(ARROW_LEFT)
                            press(TAB)





                    else:
                        press(ARROW_LEFT)
                        press(TAB)
                        write(relative['cur_age'])
                        press(ENTER)
                        press(TAB)
                        press(TAB)
                        press(TAB)
                    

                    if relative['oc_status'] == 'Yes':
                        press(ARROW_LEFT)
                        press(ARROW_LEFT)

                        press(TAB)
                        write(relative['oc_onset'])
                        press(ENTER)
                        press(TAB)
                        press(TAB)
                        press(TAB)
                    else:
                        press(ARROW_LEFT)
                        press(TAB)

                    if relative.get('brca_gene') is not None:
                        if relative['brca_gene'] == 'Negative':
                            press(ARROW_RIGHT)
                        elif relative['brca_gene'] == 'BRCA1':
                            press(ARROW_RIGHT)
                            press(ARROW_RIGHT)
                        elif relative['brca_gene'] == 'BRCA2':
                            press(ARROW_RIGHT)
                            press(ARROW_RIGHT)
                            press(ARROW_RIGHT)
                        else:
                            pass
                        
                    else:
                        pass

                elif relative['name'] == 'Maternal Aunt':
                    write('1')
                    press(ENTER)
                    press(TAB)
                    press(TAB)
                    click('GO TO DETAILS')

                    

                    to_click = str(relative['name']).upper()
                    
                    click(Link(to_click))

                    for i in range(13-int(seq_dict[relative['name']])):
                    
                        press(TAB)
                 
                    if relative['bc_status'] == 'Yes':
                        press(ARROW_LEFT)
                        press(ARROW_LEFT)
                        press(ARROW_LEFT)
                        press(TAB)
                        write(relative['cur_age'])
                        press(ENTER)
                        press(TAB)
                        press(TAB) 
                        press(TAB)
                        press(ARROW_LEFT)
                        
                        if relative['bilateral_status'] == 'Yes':
                            press(ARROW_LEFT)
                            press(ARROW_LEFT)
                            press(TAB)
                            # 
                            write(relative['bc_onset'])
                            press(ENTER)
                            press(TAB)
                            press(TAB)
                            press(TAB)

                            
                        else:
                            press(ARROW_LEFT)
                            press(TAB)





                    else:
                        press(ARROW_LEFT)
                        press(ARROW_LEFT)
                        press(TAB)
                        write(relative['cur_age'])
                        press(ENTER)
                        press(TAB)
                        press(TAB)
                        press(TAB)
                    

                    if relative['oc_status'] == 'Yes':
                        press(ARROW_LEFT)
                        press(ARROW_LEFT)

                        press(TAB)
                        write(relative['oc_onset'])
                        press(ENTER)
                        press(TAB)
                        press(TAB)
                        press(TAB)
                    else:
                        press(ARROW_LEFT)
                        press(TAB)

                    if relative.get('brca_gene') is not None:
                        if relative['brca_gene'] == 'Negative':
                            press(ARROW_RIGHT)
                        elif relative['brca_gene'] == 'BRCA1':
                            press(ARROW_RIGHT)
                            press(ARROW_RIGHT)
                        elif relative['brca_gene'] == 'BRCA2':
                            press(ARROW_RIGHT)
                            press(ARROW_RIGHT)
                            press(ARROW_RIGHT)
                        else:
                            pass
                        
                    else:
                        pass

                elif relative['name'] == 'Maternal Half Sister':
                    write('1')
                    press(ENTER)
                    press(TAB)
                    press(TAB)
                    click('GO TO DETAILS')

                    

                    to_click = str(relative['name']).upper()
                    
                    click(Link(to_click))

                    for i in range(13-int(seq_dict[relative['name']])):
                    
                        press(TAB)
                 
                    if relative['bc_status'] == 'Yes':
                        press(ARROW_LEFT)
                        press(ARROW_LEFT)
                        press(ARROW_LEFT)
                        press(TAB)
                        write(relative['cur_age'])
                        press(ENTER)
                        press(TAB)
                        press(TAB) 
                        press(TAB)
                        press(ARROW_LEFT)
                        
                        if relative['bilateral_status'] == 'Yes':
                            press(ARROW_LEFT)
                            press(ARROW_LEFT)
                            press(TAB)
                             
                            write(relative['bc_onset'])
                            press(ENTER)
                            press(TAB)
                            press(TAB)
                            press(TAB)

                            
                        else:
                            press(ARROW_LEFT)
                            press(TAB)





                    else:
                        press(ARROW_LEFT)
                        press(ARROW_LEFT)
                        press(TAB)
                        write(relative['cur_age'])
                        press(ENTER)
                        press(TAB)
                        press(TAB)
                        press(TAB)
                    

                    if relative['oc_status'] == 'Yes':
                        press(ARROW_LEFT)
                        press(ARROW_LEFT)

                        press(TAB)
                        write(relative['oc_onset'])
                        press(ENTER)
                        press(TAB)
                        press(TAB)
                        press(TAB)
                    else:
                        press(ARROW_LEFT)
                        press(TAB)

                    if relative.get('brca_gene') is not None:
                        if relative['brca_gene'] == 'Negative':
                            press(ARROW_RIGHT)
                        elif relative['brca_gene'] == 'BRCA1':
                            press(ARROW_RIGHT)
                            press(ARROW_RIGHT)
                        elif relative['brca_gene'] == 'BRCA2':
                            press(ARROW_RIGHT)
                            press(ARROW_RIGHT)
                            press(ARROW_RIGHT)
                        else:
                            pass
                        
                    else:
                        pass

                elif relative['name'] == 'Paternal Aunt':
                    write('1')
                    press(ENTER)
                    press(TAB)
                    press(TAB)
                    click('GO TO DETAILS')

                    

                    to_click = str(relative['name']).upper()
                    
                    click(Link(to_click))

                    for i in range(13-int(seq_dict[relative['name']])):
                    
                        press(TAB)
                 
                    if relative['bc_status'] == 'Yes':
                        press(ARROW_LEFT)
                        press(ARROW_LEFT)
                        press(ARROW_LEFT)
                        press(TAB)
                        write(relative['cur_age'])
                        press(ENTER)
                        press(TAB)
                        press(TAB) 
                        press(TAB)
                        press(ARROW_LEFT)
                        
                        if relative['bilateral_status'] == 'Yes':
                            press(ARROW_LEFT)
                            press(ARROW_LEFT)
                            press(TAB)
                             
                            write(relative['bc_onset'])
                            press(ENTER)
                            press(TAB)
                            press(TAB)
                            press(TAB)

                            
                        else:
                            press(ARROW_LEFT)
                            press(TAB)





                    else:
                        press(ARROW_LEFT)
                        press(ARROW_LEFT)
                        press(TAB)
                        write(relative['cur_age'])
                        press(ENTER)
                        press(TAB)
                        press(TAB)
                        press(TAB)
                    

                    if relative['oc_status'] == 'Yes':
                        press(ARROW_LEFT)
                        press(ARROW_LEFT)

                        press(TAB)
                        write(relative['oc_onset'])
                        press(ENTER)
                        press(TAB)
                        press(TAB)
                        press(TAB)
                    else:
                        press(ARROW_LEFT)
                        press(TAB)

                    if relative.get('brca_gene') is not None:
                        if relative['brca_gene'] == 'Negative':
                            press(ARROW_RIGHT)
                        elif relative['brca_gene'] == 'BRCA1':
                            press(ARROW_RIGHT)
                            press(ARROW_RIGHT)
                        elif relative['brca_gene'] == 'BRCA2':
                            press(ARROW_RIGHT)
                            press(ARROW_RIGHT)
                            press(ARROW_RIGHT)
                        else:
                            pass
                        
                    else:
                        pass

                elif relative['name'] == 'Paternal Half Sister':
                    write('1')
                    press(ENTER)
                    press(TAB)
                    press(TAB)
                    click('GO TO DETAILS')

                    

                    to_click = str(relative['name']).upper()
                    
                    click(Link(to_click))

                    for i in range(13-int(seq_dict[relative['name']])):
                    
                        press(TAB)
                 
                    if relative['bc_status'] == 'Yes':
                        press(ARROW_LEFT)
                        press(ARROW_LEFT)
                        press(ARROW_LEFT)
                        press(TAB)
                        write(relative['cur_age'])
                        press(ENTER)
                        press(TAB)
                        press(TAB) 
                        press(TAB)
                        press(ARROW_LEFT)
                        
                        if relative['bilateral_status'] == 'Yes':
                            press(ARROW_LEFT)
                            press(ARROW_LEFT)
                            press(TAB)
                            # 
                            write(relative['bc_onset'])
                            press(ENTER)
                            press(TAB)
                            press(TAB)
                            press(TAB)

                            
                        else:
                            press(ARROW_LEFT)
                            press(TAB)





                    else:
                        press(ARROW_LEFT)
                        press(ARROW_LEFT)
                        press(TAB)
                        write(relative['cur_age'])
                        press(ENTER)
                        press(TAB)
                        press(TAB)
                        press(TAB)
                    

                    if relative['oc_status'] == 'Yes':
                        press(ARROW_LEFT)
                        press(ARROW_LEFT)

                        press(TAB)
                        write(relative['oc_onset'])
                        press(ENTER)
                        press(TAB)
                        press(TAB)
                        press(TAB)
                    else:
                        press(ARROW_LEFT)
                        press(TAB)

                    if relative.get('brca_gene') is not None:
                        if relative['brca_gene'] == 'Negative':
                            press(ARROW_RIGHT)
                        elif relative['brca_gene'] == 'BRCA1':
                            press(ARROW_RIGHT)
                            press(ARROW_RIGHT)
                        elif relative['brca_gene'] == 'BRCA2':
                            press(ARROW_RIGHT)
                            press(ARROW_RIGHT)
                            press(ARROW_RIGHT)
                        else:
                            pass
                        
                    else:
                        pass

                    """

                    if relative['bc_status'] == 'Yes':
                        click(Link('Breast Cancer'))


                        if relative['bilateral_status'] == 'Yes':
                            click(Link('Bilateral'))
                            press(TAB)

                            write(relative['bc_onset'])
                            press(ENTER)
                            press(TAB)

                    
                    if relative['oc_status'] == 'Yes':
                        click(Link('Ovarian '))

                        press(TAB)
                        write(relative['oc_onset'])
                        press(ENTER)
                        press(TAB)

                    else:
                        press(TAB)


                    if relative.get('brca_gene') is not None:
                        if relative['brca_gene'] == 'Negative':
                            press(ARROW_RIGHT)
                            press(TAB)
                        elif relative['brca_gene'] == 'BRCA1':
                            press(ARROW_RIGHT)
                            press(ARROW_RIGHT)
                            press(TAB)
                        elif relative['brca_gene'] == 'BRCA2':
                            press(ARROW_RIGHT)
                            press(ARROW_RIGHT)
                            press(ARROW_RIGHT)
                            press(TAB)
                        else:
                            press(TAB)
                        
                    else:
                        press(TAB)

                    write(relative['cur_age'])
                    press(ENTER)

                """

                else:
                    continue
            
            
        click('CALCULATE RISK')
        scroll_down(num_pixels=200)


        time.sleep(3)

        self.driver.save_screenshot('screenshot.png')



        html=self.driver.page_source
        with open('ser.txt','w',encoding='utf-8') as file:
            file.write(str(html))
            file.close()


        soup=BeautifulSoup(html,'html.parser')




        tab = soup.find("div",{"class":"textbox"})
        risk10 = tab.find("span",{"id":"text_personal_10y_risk"})
        risklt = tab.find("span",{"id":"text_personal_lt_risk"})
        risk10 = str(risk10.text)
        risklt = str(risklt.text)

        file = open("AI External-Outputs/path_info.txt","r")
        for lines in file.read().splitlines():
            if lines[0] == "T":
                test_filepath = lines[1:]
        file.close() 
        
        t_filename = Path(test_filepath).stem

        print("Genetic Risk = 10 Year Risk = {} Lifetime Risk = {}".format(risk10,risklt))

        df = pd.read_csv("AI External-Outputs/Prediction_output_{}.csv".format(t_filename))
        dfn = df.iloc[self.case_iter,10] = risk10
        df_n2 = df.iloc[self.case_iter,11] = risklt
        df.to_csv("AI External-Outputs/Prediction_output_{}.csv".format(t_filename),index=False)
        # self.driver.close()
        conn = http.client.HTTPConnection(self.driver.service.service_url.split("//")[1])
        conn.request("GET", "/shutdown")
        conn.close()
        del self.driver

        """
        

        workbook = load_workbook(("AI External-Outputs/Prediction_output_{}.xlsx".format(t_filename)))
        sheet = workbook.active

        sheet.cell(row=1 + self.case_iter, column=12).value = risk10
        sheet.cell(row=1 + self.case_iter, column=13).value = risklt

        workbook.save(("AI External-Outputs/Prediction_output_{}.xlsx".format(t_filename)))


        """


      





    






        


        




    

        
            



  