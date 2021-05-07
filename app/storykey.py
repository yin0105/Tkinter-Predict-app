from os import curdir
from tkinter import *
import os
import tkinter.filedialog
import numpy as np
from styleframe import StyleFrame, utils
import math
import pandas as pd
import numpy as np

import csv
import matplotlib
from tkmagicgrid import *
from PIL import ImageTk, Image
from tkinter import ttk
import uuid
from itertools import count
from tkinter.tix import ScrolledWindow
import json
import requests
from styleframe import StyleFrame, utils
import random
import pandas as pd
import numpy as np
from sentence_splitter import SentenceSplitter, split_text_into_sentences
from gingerit.gingerit import GingerIt
from os import curdir
import collections
import tkinter.font as tkFont
from openpyxl import Workbook
from openpyxl import load_workbook
from pathlib import Path
class storykey():

    def __init__(self):
        self.name = "Mr. Smith"

        self.age = 32
        self.accompanied_by = "Mr. John"
        self.location = "Brighton High"
        self.category_logic = "Category-1,Category-2"
        self.struct_dict = {
                          "address_logic" : self.get_address_logic,
                          "report_logic" : self.get_report_logic,
                          "data" : self.get_data_logic,
                          "name" : self.get_name,
                          "age" : self.get_age,
                          "accompanied_by" : self.get_accompanied_by,
                          "location_data" : self.get_location,
                          "gender_pronoun" : self.get_pronoun,
                          "category_logic" : self.get_category_logic,
                          "also-1" : self.get_also1_logic,
                          "also-2" : self.get_also2_logic,
                          "category-1" : self.get_category_1,
                          "category-2" : self.get_category_2,
                          "category-3" : self.get_category_3,
                          "category-4" : self.get_category_4,
                          "category-5" : self.get_category_5,
                          "category-6" : self.get_category_6,
                          "category-7" : self.get_category_7,
                          "category-8" : self.get_category_8,
                          "category-9" : self.get_category_9
                          }

        self.idx = 0

        self.structure_logic_sheet = os.path.join(curdir,"Input Sheets","structure_logic_sheet.xlsx")
        self.story_logic_sheet = os.path.join(curdir,"Input Sheets","story_logic_sheet.xlsx")



    def display_output(self,sent):
        file = open("AI External-Outputs/path_info.txt","r")
        for lines in file.read().splitlines():
            if lines[0] == "T":
                test_filepath = lines[1:]
        file.close() 
        
        t_filename = Path(test_filepath).stem
        df = pd.read_csv("AI External-Outputs/Prediction_output_{}.csv".format(t_filename))
        df_n = df.iloc[self.case_iter,7] = sent
        #df.iloc[1+self.case_iter,13] = risklt
        df.to_csv("AI External-Outputs/Prediction_output_{}.csv".format(t_filename),index=False)


        """

        workbook = load_workbook(("AI External-Outputs/Prediction_output_{}.xlsx".format(t_filename)))
        sheet = workbook.active

        print("Story is = {}".format(sent))

        sheet.cell(row=1 + self.case_iter, column=10).value = sent
    

        workbook.save(("AI External-Outputs/Prediction_output_{}.xlsx".format(t_filename)))



        """


    def del_and_open(self):
        self.test_sent = ""
        
        print("Succesfully Scanned")
        self.test_sheet = self.test_sheet.rstrip()
        self.test_sheet = self.test_sheet.lstrip()
        self.story_logic_sheet = self.story_logic_sheet.rstrip()
        self.story_logic_sheet = self.story_logic_sheet.lstrip()

        df_attempt = self.get_attempt_sheet()
        self.df_attempt = self.pad_array(df_attempt,self.row_lim+1)
        self.get_total_df()
        self.fetch_data()

        clb_sheet = self.get_clubbing_sheet()
        self.clb_sheet = self.pad_array(clb_sheet,self.row_lim+1)
        remark_sheet = self.get_remark_sheet()
        self.remark_sheet = self.pad_array(remark_sheet,self.row_lim+1)
        dependency_sheet = self.get_dependency_sheet()
        self.dependency_sheet = self.pad_array(dependency_sheet,self.row_lim+1)
        category_sheet = self.get_category_sheet()
        self.category_sheet = self.pad_array(category_sheet,self.row_lim+1)
        data_logic_sheet = self.get_data_logic_sheet()
        self.data_logic_sheet = self.pad_array(data_logic_sheet,self.row_lim+1)
        self.get_os()


        total_sent = self.os+ "." +"\n"+ " "+self.get_data_dict()
        total_sent = total_sent.replace(' .','')
        total_sent = total_sent.replace("\n\n. ","\n\n")




        self.display_output(total_sent)

        

    def get_case_num(self,case_num):
      self.case_num = case_num
      self.del_and_open()
   
    def get_name(self):
      return self.name
  
    def get_age(self):
      return self.age

    def get_accompanied_by(self):
      return self.accompanied_by

    def get_location(self):
      return self.location

    def get_category_logic(self):
      self.category_logic = []
      tot_cat = [self.category_1,self.category_2,self.category_3,self.category_4,self.category_5,self.category_6,self.category_7,self.category_8,self.category_9]
      for cat in tot_cat:
        if cat!=cat:
          pass
        else:
          self.category_logic.append(cat)
      return ','.join(self.category_logic)


    def get_category_1(self):
      return self.category_1
    
    def get_category_2(self):
      return self.category_2

    def get_category_3(self):
      return self.category_3

    def get_category_4(self):
      return self.category_4

    def get_category_5(self):
      return self.category_5
    def get_category_6(self):
      return self.category_6

    def get_category_7(self):
      return self.category_7

    def get_category_8(self):
      return self.category_8

    def get_category_9(self):
      return self.category_9




      
    def fetch_data(self):
      self.df_story_logic = pd.read_excel(self.story_logic_sheet)
      self.df_test_sheet = pd.read_excel(self.test_sheet)

      self.df_structure_logic = pd.read_excel(self.structure_logic_sheet)

      self.category_1 = self.df_structure_logic['Category-1'].values[0]
      self.category_2 = self.df_structure_logic['Category-2'].values[0]
      self.category_3 = self.df_structure_logic['Category-3'].values[0]
      self.category_4 = self.df_structure_logic['Category-4'].values[0]
      self.category_5 = self.df_structure_logic['Category-5'].values[0]
      self.category_6 = self.df_structure_logic['Category-6'].values[0]
      self.category_7 = self.df_structure_logic['Category-7'].values[0]
      self.category_8 = self.df_structure_logic['Category-8'].values[0]
      self.category_9 = self.df_structure_logic['Category-9'].values[0]


      self.clb_choice = self.df_structure_logic['Clubbing logic'].values

      self.rpt_choice = self.df_structure_logic['Report logic'].values

      self.addrs_choice = self.df_structure_logic['Address logic'].values
      #self.also_choice = df_structure_logic['Also logic']
      self.os_logic = self.df_structure_logic['OS logic'].values[0]
      self.location = self.df_structure_logic['Location'].values[0]
      self.also1_choice = self.df_structure_logic['Also-1 logic'].values
      self.also2_choice = self.df_structure_logic['Also-2 logic'].values

    def get_pronoun(self):
      if self.gender == "Female":
        return 'She'
      else:
        return 'He'

    

    def get_user_dict(self):
      df_test_sheet= pd.read_excel(self.test_sheet)
      chl = 0
      chlt = 0
      for inx,rows in df_test_sheet.iterrows():
        if rows['Sub-Feature'] == "external factor":
          chl = inx
      
      for inx,rows in df_test_sheet.iterrows():
        if rows['Sub-Feature'] == "endFeatures":
          chlt = inx
      
      df2 = df_test_sheet[self.case_num]
      self.name = df2.iloc[1:2].values[0]
      self.gender = df2.iloc[2:3].values[0]
      self.accompanied_by = df2.iloc[3:4].values[0]
      self.ethinicity = df2.iloc[4:5].values[0]
      self.age = df2.iloc[5:6].values[0]
      self.date = df2.iloc[6:7].values[0]
      self.issues = df2.iloc[7:8].values[0]
      
      
      df2 = df2.fillna(0)
      print(chl)
      self.df_test_sheet = df2.iloc[chl:chlt].values
      self.df_test_total = df2.iloc[chl:chlt].values
      

      self.row_lim = len(df_test_sheet)
      df_attempt = self.df_test_sheet
      df_attempt = np.where(df_attempt == 0,0,1)
      return df_attempt
      

      

    


    def get_os(self):
      """os_list = []
      for os_logic in self.os_logic:
        os_list.append(self.get_logic_sent(os_logic))
      print(os_list)
      """
      self.os = str(self.get_logic_sent(self.os_logic))

    def get_attempt_sheet(self):
      df_test_sheet= pd.read_excel(self.test_sheet)
      chl = 0
      chlt = 0
      for inx,rows in df_test_sheet.iterrows():
        if rows['Sub-Feature'] == "external factor":
          chl = inx
      
      for inx,rows in df_test_sheet.iterrows():
        if rows['Sub-Feature'] == "DropDowns":
          chlt = inx
      
      df2 = df_test_sheet[self.case_num]
      self.name = df2.iloc[1:2].values[0]
      self.gender = df2.iloc[2:3].values[0]
      self.accompanied_by = df2.iloc[3:4].values[0]
      self.ethinicity = df2.iloc[4:5].values[0]
      self.age = df2.iloc[5:6].values[0]
      self.date = df2.iloc[6:7].values[0]
      self.issues = df2.iloc[7:8].values[0]
      
      
      df2 = df2.fillna(0)
     
      self.df_test_sheet = df2.iloc[chl:chlt].values
      self.df_test_total = df2.iloc[chl:chlt].values
      

      self.row_lim = len(df_test_sheet)
      df_attempt = self.df_test_sheet
      df_attempt = np.where(df_attempt == 0,0,1)
      return df_attempt
      

      

    


    def make_me_plural(self,sent):
      sent = self.get_logic_sent(sent) + 's'
      return sent

    def get_dependency_logic(self,data_dict):

      if data_dict.get('B3-c') is not None or data_dict.get('B3-d') is not None:
        if data_dict.get('A-a') is not None:
          
          data_dict['A-a'] = self.make_me_plural(data_dict['A-a'])
        elif data_dict.get('A-b') is not None:
          
          data_dict['A-b'] = self.make_me_plural(data_dict['A-b'])

        else:
          pass


      else:
        pass

      return data_dict
        


    def get_total_df(self):
      df_story_logic = pd.read_excel(self.story_logic_sheet)

      self.total_df = df_story_logic.iloc[:self.row_lim,:]

    def get_category_sheet(self):
      category_sheet = self.total_df['Category'].values
      uniqueList = []
      for elem in category_sheet:
          if str(elem) not in uniqueList:
              uniqueList.append(str(elem))
      self.category_logic = ','.join(uniqueList)
      #category_sheet = list(map(int, category_sheet))
      return category_sheet

    def get_category_value(self):


      return str(self.category_sheet[self.idx])


    def get_data_logic_sheet(self):
      data_logic_sheet = self.df_test_total
      return data_logic_sheet


    def get_clubbing_sheet(self):
      df = self.total_df.fillna(0)
      df = df["Clubbing"].values
      clb_sheet = list(map(int, df))
      return clb_sheet

    def get_logic_sent(self,text):
      keys = self.struct_dict.keys()
      text = str(text)
      for key in keys:

        if key in text:
          if self.struct_dict.get(key)() is not None:
            text = text.replace('<'+key+'>',str(self.struct_dict.get(key)()))


      return text


    def remove_nan(self,arr):
      new_arr = []
      for br in arr:
        if br!=br or br=="nan":
          pass
        else:
          new_arr.append(br)
      return new_arr

    def get_data_logic(self):

      return self.data_logic_sheet[self.idx]



    def get_report_logic(self):
      new_rpt_choice = []
      for ch in self.rpt_choice:
        if ch is not None:
          new_rpt_choice.append(self.get_logic_sent(ch))
      new_rept_choice = self.remove_nan(new_rpt_choice)

  
      return str(random.choice(new_rept_choice))


    def get_also1_logic(self):
      new_addrs_choice = []
      for ch in self.also1_choice:
        if ch is not None:
          new_addrs_choice.append(self.get_logic_sent(ch))


      new_ads_choice = self.remove_nan(new_addrs_choice)



      return str(random.choice(new_ads_choice))

    def get_also2_logic(self):
      new_addrs_choice = []
      for ch in self.also2_choice:
        if ch is not None:
          new_addrs_choice.append(self.get_logic_sent(ch))


      new_ads_choice = self.remove_nan(new_addrs_choice)



      return str(random.choice(new_ads_choice))

      



    def get_address_logic(self):
      new_addrs_choice = []
      for ch in self.addrs_choice:
        if ch is not None:
          new_addrs_choice.append(self.get_logic_sent(ch))


      new_ads_choice = self.remove_nan(new_addrs_choice)



      return str(random.choice(new_ads_choice))


    def get_clb_logic(self):
      new_clb_choice = []
      for ch in self.clb_choice:
        if ch is None:
          pass
        else:
          new_clb_choice.append(self.get_logic_sent(ch))


      new_clbt_choice = self.remove_nan(new_clb_choice)


    
      return str(random.choice(new_clbt_choice))

    def get_alphanumeric(self,key):
      return any(char.isdigit() for char in key)


    def pad_array(self,d,length):
      #print(len(np.pad(d, (0,(length - len(d)%length)), 'constant')))
      return np.pad(d, (0,(length - len(d)%length)), 'constant')

      

    def get_alpha(self):
      pass

    def get_dependency_sheet(self):
      def get_standard_matrix():
        filepath = self.story_logic_sheet

        sf = StyleFrame.read_excel(filepath , read_style=True, use_openpyxl_styles=False)




        def only_cells_with_red_text(cell):

            
            if cell.style.bg_color in {utils.colors.red, 'FFFF0000'}:
                return 120
        


        
        sf_2 = StyleFrame(sf.applymap(only_cells_with_red_text))





        #print(qualifying_dict)


        sf_2.to_excel().save()
        df = pd.read_excel(os.path.join(curdir,"AI Internal-Outputs",'output.xlsx'))


        df = df.iloc[:self.row_lim,1]
        #print(df)

        standard_matrix = df.values


        
        return standard_matrix
      
      sm = get_standard_matrix()
      sm = np.where(sm == 120,1,0)

      return sm

    def get_remark_sheet(self):
      def get_standard_matrix():
        filepath = self.story_logic_sheet
        print("#" * 50)
        print("filepath = ", filepath)

        sf = StyleFrame.read_excel(filepath , read_style=True, use_openpyxl_styles=False)

        def only_cells_with_red_text(cell):
            if cell.style.bg_color in {utils.colors.red, 'FFFF0000'}:
                return 120
        
        sf_2 = StyleFrame(sf.applymap(only_cells_with_red_text))





        #print(qualifying_dict)


        sf_2.to_excel().save()
        df = pd.read_excel(os.path.join(curdir,"AI Internal-Outputs",'output.xlsx'))


        df = df.iloc[:,3]
        #print(df)

        standard_matrix = df.values


        
        return standard_matrix
      
      sm = get_standard_matrix()
      sm = np.where(sm == 120,1,0)

      return sm
                              





    def get_struct_base(self,sre):
      sre = sre.replace("<","")
      sre = sre.split(">")
      structure = []
      for sr in sre:
        if sr == "":
          pass
        else:
          structure.append(sr)

      return structure



    def get_clubbing_logic(self,list_of_keys):
      
      

      clb_sent =  str(self.get_clb_logic()) + " " +str("or".join(list_of_keys))
      return clb_sent
      
    def get_data_dict(self):
      clb_sent = []
      data_dict = {}

      total_sent = ""
      self.idx  = 1
      back_category_value = self.get_category_value()
      for self.idx,row in self.total_df.iterrows():
        if self.get_category_value()!=None:
          if back_category_value == self.get_category_value():
        
            
            if self.remark_sheet[self.idx] == 0:
              if self.df_attempt[self.idx] == 1:
                if self.clb_sheet[self.idx] == 1:
                  if row['AN-Data dictionary']!=row['AN-Data dictionary']:
                    pass
                  else:
                    clb_sent.append(self.get_logic_sent(str(row['AN-Data dictionary'])))
                  
                else:

                  if row.Sequence!=row.Sequence:
                    pass
                  else:

                    if self.get_alphanumeric(str(row.Sequence)):
      
                      
                        data_dict[row.Sequence] = self.get_logic_sent(str(row['AN-Data dictionary']))
                    
                  
                


                        

                        
                      #print(row.Sequence)
                      
                    else:
                      data_dict[row.Sequence] = "."+self.get_logic_sent(str(row['A-Sentence']))+"."
      
                    

              
              else:
                pass
            else:
              pass
          else:
            back_category_value = self.get_category_value()
            

            for key,value in data_dict.items():
              data_dict[key] = self.get_logic_sent(str(value))
            data_dict = self.get_dependency_logic(data_dict)
            

            keys_values = data_dict.items()

            data_dict = {str(key): str(value) for key, value in keys_values}

            
            







            total_sent = total_sent + self.get_my_sentence(data_dict,clb_sent) + "\n\n"



            clb_sent = []
            data_dict = {}
    



      for key,value in data_dict.items():
        data_dict[key] = self.get_logic_sent(str(value))
      data_dict = self.get_dependency_logic(data_dict)
      
      keys_values = data_dict.items()

      data_dict = {str(key): str(value) for key, value in keys_values}

      


      total_sent = total_sent + self.get_my_sentence(data_dict,clb_sent) 

            




      return total_sent



    def form_sentence(self,data_dict,clb_sent):
      with open(os.path.join(curdir,'Logic Container',"sentence_structure.txt")) as tweetfile:
        sentstruct = json.load(tweetfile)
      sent = ""
      for i in range(1,5):
        if sentstruct['address_logic'] == i:
          sent = sent + "." + " " + self.get_address_logic() + " "
        elif sentstruct['report_logic'] == i:
          sent = sent + self.get_report_logic() + " "

        elif sentstruct['data_logic'] == i:
          for key,value in data_dict.items():
            sent = sent + str(value) + " "
        elif sentstruct['clubbing_logic'] == i:
          sent = sent + "." + clb_sent
        else:
          pass




      return sent

    def gramarize(self,sent):

      f = open('Value-json/logic_activation.json') 
      activation = json.load(f)
      if activation['grammar_logic'] == "active":
        test_str = sent
        splitter = SentenceSplitter(language='en')
        sente = splitter.split(text=test_str)
        gram_sent = []

        for sent in sente:
          parser = GingerIt()

          output = parser.parse(sent)
          output_1 = (output.get("result"))
          output_1 = output_1 
          gram_sent.append(output_1)

        f_output = ' '.join(gram_sent)


        if f_output[-1] == '.' and f_output[-2] == '.':
          f_output = f_output[:-2]




        
        f_output = f_output + '.'

        f_output = self.remove_trailing_dots(f_output)
        f_output = f_output.replace('..','.')

        return f_output


      else:
        return sent


    def remove_trailing_dots(self,f_output):
      if f_output[-1] == '.' and f_output[-2] == '.':
        f_output = f_output[:-2]
      f_output = f_output + '.'
      return f_output


    def get_my_sentence(self,data_dict,clb_sent):
      data_dict = collections.OrderedDict(sorted(data_dict.items()))
      if len(clb_sent) > 0:
        clb_sent = self.get_clubbing_logic(clb_sent)
      else:
        clb_sent = ""
      sent = self.form_sentence(data_dict,clb_sent)
      sent = sent.replace("nan","")
      
      return str(self.gramarize(sent))


    def execute(self,test_sheet,case_num,case_iter):
      self.case_iter = case_iter
      self.test_sheet = test_sheet
      self.get_case_num(case_num)
        #self.read_google_sheet()
